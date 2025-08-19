import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO

def analisar_produtos(arquivo_excel):
    """
    Esta função realiza uma análise dos produtos listados em um arquivo Excel, calcula um score ponderado
    considerando nota, quantidade de avaliações e preço, gera um gráfico de Preço x Nota, identifica os
    3 melhores produtos (Top 3) e adiciona os resultados em uma aba "Analise" na própria planilha.

    Parâmetros:
    arquivo_excel (str): Caminho para o arquivo Excel contendo os dados dos produtos. Espera-se que tenha
        colunas como "TITULO", "PRECO", "URL", "NOTA_AVAL" e "QTD_AVAL".

    Retorna:
    None: A função salva as alterações diretamente no arquivo Excel fornecido, criando a aba "Analise"
        com estatísticas gerais, tabela do Top 3 e gráfico.
    """

    # --- Leitura da planilha ---
    df = pd.read_excel(arquivo_excel)
    
    # --- Tratamento da coluna PRECO ---
    df['PRECO'] = (
        df['PRECO']
        .astype(str)
        .str.replace(r'R\$', '', regex=True)
        .str.replace('.', '', regex=False)
        .str.replace(',', '.', regex=False)
        .astype(float)
    )
    
    # --- Normalização dos fatores para SCORE ---
    nota_norm = df['NOTA_AVAL'] / 5  # entre 0 e 1
    qtd_av_norm = df['QTD_AVAL'] / df['QTD_AVAL'].max()  # 0 a 1
    preco_norm = 1 - (df['PRECO'] - df['PRECO'].min()) / (df['PRECO'].max() - df['PRECO'].min()) 
    
    # --- Cálculo do SCORE final ---
    # ponderação: nota 50%, quantidade 30%, preço 20%
    df['SCORE'] = (nota_norm * 0.5 + qtd_av_norm * 0.3 + preco_norm * 0.2) * 10  # escala 0-10
    
    # Top 3 produtos
    top3 = df.sort_values(by='SCORE', ascending=False).head(3).reset_index(drop=True)
    
    # --- Gráfico Preço x Nota ---
    plt.figure(figsize=(8,6))
    plt.scatter(df['PRECO'], df['NOTA_AVAL'], s=50, alpha=0.7, c="blue", label="Produtos")
    plt.scatter(top3['PRECO'], top3['NOTA_AVAL'], s=120, c="red", marker="*", label="Top 3")
    
    # Coloca apenas o ranking no gráfico
    for i, row in top3.iterrows():
        plt.text(row['PRECO'], row['NOTA_AVAL']+0.05, f"{i+1}º", 
                 fontsize=10, color="red", ha="center", weight="bold")
    
    plt.xlabel('Preço (R$)')
    plt.ylabel('Nota de Avaliação')
    plt.title('Análise Geral: Preço x Avaliação')
    plt.legend()
    plt.grid(True, linestyle="--", alpha=0.6)
    plt.tight_layout()
    
    # Salvar gráfico em buffer
    grafico_buffer = BytesIO()
    plt.savefig(grafico_buffer, format='png')
    plt.close()
    grafico_buffer.seek(0)
    
    # --- Abrir planilha ---
    wb = load_workbook(arquivo_excel)
    
    # Cria ou substitui aba "Analise"
    if "Analise" in wb.sheetnames:
        ws = wb["Analise"]
        wb.remove(ws)
    ws = wb.create_sheet("Analise")
    
    # Estatísticas gerais
    ws["A1"] = "📊 Estatísticas Gerais"
    ws["A2"] = f"Preço médio: R$ {df['PRECO'].mean():.2f}"
    ws["A3"] = f"Preço mínimo: R$ {df['PRECO'].min():.2f}"
    ws["A4"] = f"Preço máximo: R$ {df['PRECO'].max():.2f}"
    ws["A5"] = f"Média de avaliações: {df['NOTA_AVAL'].mean():.2f}"
    
    # Tabela Top 3
    ws["A7"] = "🎯 Top 3 Produtos Recomendados"
    colunas = ["Ranking", "Título", "Preço (R$)", "Nota", "Score", "Link"]
    for j, col in enumerate(colunas, 1):
        ws.cell(row=8, column=j, value=col)
    
    for i, row in top3.iterrows():
        ws.cell(row=9+i, column=1, value=f"{i+1}º")
        ws.cell(row=9+i, column=2, value=row['TITULO'])
        ws.cell(row=9+i, column=3, value=round(row['PRECO'], 2))
        ws.cell(row=9+i, column=4, value=row['NOTA_AVAL'])
        ws.cell(row=9+i, column=5, value=round(row['SCORE'], 2))
        ws.cell(row=9+i, column=6, value=row['URL'])
    
    # Inserir gráfico
    img_graf = XLImage(grafico_buffer)
    img_graf.width = 600
    img_graf.height = 400
    ws.add_image(img_graf, "G2")
    
    # Salvar planilha
    wb.save(arquivo_excel)
    print("✅ Análise concluída! Aba 'Analise' criada com estatísticas, tabela de ranking e gráfico.")
